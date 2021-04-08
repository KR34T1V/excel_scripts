from search import search
from add import add
from src import source
from openpyxl import load_workbook

import sys

def main(to, fra, column, src):
  print('Preparing Workbooks')
  wb_src = load_workbook(src)
  wb_fra = load_workbook(fra, keep_vba=True)
  wb_to = load_workbook(to)
  print('Workbooks Loaded')
  if (not to or not fra or not column or not src):
    return
  i = 1
  while (1):
    print('('+ str(i) +')')
    email = source(wb_src, i)
    if (email == -1):
      break
    if (email != None):
      add(wb_to, wb_fra, column, email)
    i = i + 1
  wb_to.save(to)
  print('Saved\n')
  print('Program Complete!\n')


if (len(sys.argv) >= 3):
  main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
else:
  print("Please provide the following arguments\n to, from, column, email_src\n")